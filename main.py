from fastapi import FastAPI, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from datetime import date
import os
import json
import shutil
from openpyxl import load_workbook
import logging
from bs4 import BeautifulSoup
import re
import hashlib
import time

from pdf_generator import modify_pdf  # Import the modify_pdf function
from looping_mechanism import update_summary

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

USERPROFILE = os.environ.get("USERPROFILE")
one_drive_base = os.path.join(USERPROFILE, "OneDrive", "Desktop", "kaddar tarvaux beta version")
local_base = os.path.join(USERPROFILE, "Desktop", "kaddar tarvaux beta version")

if os.path.exists(one_drive_base):
    BASE_DIR = one_drive_base
else:
    BASE_DIR = local_base

DATABASE_DIR = os.path.join(BASE_DIR, "Database")
COUNTER_FILE = os.path.join(DATABASE_DIR, "counters.json")
EXCEL_TEMPLATE = os.path.join(DATABASE_DIR, "MODEL_DES_CHARGES.xlsx")
PDF_TEMPLATE = os.path.join(DATABASE_DIR, "template.pdf")

SUMMARY_FILE = os.path.join(DATABASE_DIR, "SUMMARYPRO.txt")
HASH_FILE = os.path.join(DATABASE_DIR, "summary_hashes.json")

os.makedirs(DATABASE_DIR, exist_ok=True)
if not os.path.exists(COUNTER_FILE):
    with open(COUNTER_FILE, 'w') as f:
        json.dump({"MP": 0, "BC": 0}, f)

def validate_submission(html_content: str) -> bool:
    return 'Lettre de confirmation.pdf' in html_content

def get_next_ref(doc_type: str) -> int:
    try:
        with open(COUNTER_FILE, 'r+') as f:
            counters = json.load(f)
            counters[doc_type] += 1
            f.seek(0)
            json.dump(counters, f)
            return counters[doc_type]
    except Exception as e:
        logger.error(f"Counter error: {str(e)}")
        raise HTTPException(500, "Erreur de compteur")

def extract_data(html: str) -> dict:
    soup = BeautifulSoup(html, 'html.parser')
    data = {
        "object": "X",
        "company": "X",
        "acheteur": "X",
        "location": "X",
        "total_ht": "0",
        "total_tva": "0",
        "total_ttc": "0",
    }

    try:
        acheteur_tag = soup.find('span', string='Acheteur public')
        if acheteur_tag:
            data["acheteur"] = acheteur_tag.find_next('span').get_text(strip=True)

        # New extraction logic for company name using the selector #dropdownMenuButton1
        company_elem = soup.find(id="dropdownMenuButton1")
        if company_elem:
            # Remove any <i> element (icon) if present
            icon = company_elem.find("i")
            if icon:
                icon.extract()
            data["company"] = company_elem.get_text(strip=True)

        obj_section = soup.find('span', class_='text-uppercase', string='Objet')
        if obj_section:
            data["object"] = obj_section.find_next('span', class_='text-black').get_text(strip=True)

        location_tag = soup.find('span', string=re.compile(r'Lieu d[\'’]exécution', re.IGNORECASE))
        if location_tag:
            data["location"] = location_tag.find_next('span').get_text(strip=True)

        totals = {
            "Total HT": "total_ht",
            "Total TVA": "total_tva",
            "Total TTC": "total_ttc"
        }
        for label, key in totals.items():
            elem = soup.find('p', string=label)
            if elem:
                value = elem.find_next('p').get_text(strip=True)
                data[key] = re.sub(r'[^\d]', '', value.split(',')[0])

        return data

    except Exception as e:
        logger.error(f"Extraction error: {str(e)}")
        return data

@app.post("/generate-reference/")
async def generate_reference(
    filename: str = Form(...),
    document_date: date = Form(...),
    document_type: str = Form(...),
    gerant: str = Form(...),
    html_content: str = Form(...)
):
    if not validate_submission(html_content):
        raise HTTPException(400, "Vous n'avez pas soumis ou le marché n'a pas encore été attribué")

    try:
        doc_type_abbr = "MP" if "Marché public" in document_type else "BC"
        ref_num = get_next_ref(doc_type_abbr)
        mm_yy = f"{document_date.month}-{document_date.year}"
        ref_id = f"{mm_yy}_{doc_type_abbr}{ref_num:03d}_{filename}"

        extracted = extract_data(html_content)

        doc_paths = []
        for folder in ["CHARGES", "Facturation", os.path.join("Archives", "Bon - Historique")]:
            full_path = os.path.join(BASE_DIR, folder, mm_yy.replace('-', '_'))
            os.makedirs(full_path, exist_ok=True)

            if "CHARGES" in folder:
                excel_path = os.path.join(full_path, f"{ref_id}.xlsx")
                shutil.copy(EXCEL_TEMPLATE, excel_path)
                wb = load_workbook(excel_path)
                ws = wb.active
                ws['A2'] = extracted["company"]
                ws['A3'] = extracted["acheteur"]
                ws['C3'] = extracted["object"]
                ws['F2'] = extracted["location"]
                ws['I2'] = int(extracted["total_ht"])
                ws['I4'] = int(extracted["total_tva"])
                ws['I6'] = int(extracted["total_ttc"])
                ws['A57'] = gerant
                wb.save(excel_path)
                doc_paths.append(excel_path)

                update_summary(excel_path)

            elif "Facturation" in folder:
                company_first = extracted["company"].split()[0].lower() if extracted["company"] else "unknown"
                first_pdf_ref = f"{mm_yy}_f_{company_first}{ref_num:03d}_{filename}"
                second_pdf_ref = f"{mm_yy}_bl_{company_first}P{ref_num:03d}_{filename}"
                pdf_path_first = os.path.join(full_path, f"{first_pdf_ref}.pdf")
                pdf_path_second = os.path.join(full_path, f"{second_pdf_ref}.pdf")
                modify_pdf(html_content, pdf_path_first)
                modify_pdf(html_content, pdf_path_second)
                doc_paths.append(pdf_path_first)
                doc_paths.append(pdf_path_second)

            else:
                archive_folder = os.path.join(full_path, ref_id)
                os.makedirs(archive_folder, exist_ok=True)
                doc_paths.append(archive_folder)

        return {
            "status": "success",
            "reference": ref_id,
            "created_files": doc_paths,
            "extracted_data": {
                **extracted,
                "gerant": gerant
            }
        }

    except Exception as e:
        logger.error(f"Generation failed: {str(e)}")
        raise HTTPException(500, f"Erreur de génération: {str(e)}")

@app.get("/health")
def health_check():
    return {"status": "ok"}
