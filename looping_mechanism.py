import os
import json
import hashlib
import time
import logging
import threading
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

# --------------------------
# Determine Base Directory
# --------------------------
# Define the potential OneDrive path and the local fallback path.
onedrive_base = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "kaddar tarvaux beta version")
local_base = os.path.join(os.path.expanduser("~"), "kaddar tarvaux beta version")

if os.path.exists(onedrive_base):
    BASE_DIR = onedrive_base
else:
    BASE_DIR = local_base

# --------------------------
# Configuration
# --------------------------
CHARGES_DIR = os.path.join(BASE_DIR, "CHARGES")  # Directory to monitor
DATABASE_DIR = os.path.join(BASE_DIR, "Database")
SUMMARY_FILE = os.path.join(DATABASE_DIR, "SUMMARYPRO.txt")
HASH_FILE = os.path.join(DATABASE_DIR, "summary_hashes.json")

# Ensure directories exist
os.makedirs(CHARGES_DIR, exist_ok=True)
os.makedirs(DATABASE_DIR, exist_ok=True)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Global dictionary for debouncing modifications.
modification_timers = {}

# --------------------------
# Initialization
# --------------------------
def initialize_summary_file():
    """Ensure the summary file exists with the proper header."""
    header = ("reference|lieu|TOTAL DES CHARGES NON JUSTIFIÉES|main_oeuvre|total_bons|"
              "total_fournisseur|tva|ttc|ht|benefice|tva_10|etat_payment\n")
    if not os.path.exists(SUMMARY_FILE):
        with open(SUMMARY_FILE, 'w') as f:
            f.write(header)
        logger.info(f"Created summary file: {SUMMARY_FILE}")

# --------------------------
# Helper Functions
# --------------------------
def get_excel_hash(file_path: str) -> str:
    """Generate a hash for the contents of an Excel file."""
    with open(file_path, 'rb') as f:
        return hashlib.md5(f.read()).hexdigest()

def get_merged_cell_value(ws, cell_coord: str):
    """
    If the cell at cell_coord is part of a merged range,
    return the value from the top-left cell of that merged range.
    Otherwise, return the cell's own value.
    """
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return ws[cell_coord].value

def extract_excel_data(file_path: str) -> dict:
    """Extract required data from an Excel file."""
    try:
        wb = load_workbook(file_path, data_only=True)
        # Always use the sheet "Feuil1" if it exists.
        if "Feuil1" in wb.sheetnames:
            ws = wb["Feuil1"]
        else:
            ws = wb.active

        def get_number(cell):
            value = ws[cell].value
            try:
                return float(value) if value is not None else 0.0
            except Exception:
                return 0.0

        def get_etat_payment():
            """
            Helper-cell logic:
            Check helper cells C59, G59, and J59 (from left to right).
            Use the rightmost helper cell that is not empty.
            Then, extract the content from its corresponding content cell
            based on this mapping:
                - C59 -> A59
                - G59 -> D59   (if the desired cell E59 is merged, D59 is used as the primer)
                - J59 -> H59
            If none of the helper cells has a value, return "PAS ENCORE".
            """
            helper_to_content = {
                "C59": "A59",
                "G59": "D59",
                "J59": "H59"
            }
            chosen_helper = None
            for helper in ["C59", "G59", "J59"]:
                if ws[helper].value not in (None, ""):
                    chosen_helper = helper
            if chosen_helper:
                content_cell = helper_to_content[chosen_helper]
                return get_merged_cell_value(ws, content_cell)
            return "PAS ENCORE"

        etat_payment = get_etat_payment()

        return {
            'reference': os.path.basename(file_path).split('.')[0],
            'lieu': ws['F2'].value or "N/A",
            'charges_non_justifiees': get_number('G43'),
            'main_oeuvre': get_number('G45'),
            'total_bons': get_number('G47'),
            'total_fournisseur': get_number('G49'),
            'tva': get_number('C53'),
            'ttc': get_number('C57'),
            'ht': get_number('H53'),
            'benefice': get_number('H55'),
            'tva_10': get_number('H57'),
            'etat_payment': etat_payment
        }
    except Exception as e:
        logger.error(f"Error processing {file_path}: {str(e)}")
        return {}

def get_all_excel_files() -> list:
    """Recursively get all Excel (.xlsx) files in CHARGES_DIR."""
    excel_files = []
    for root, dirs, files in os.walk(CHARGES_DIR):
        for file in files:
            if file.endswith(".xlsx"):
                excel_files.append(os.path.join(root, file))
    return excel_files

# --------------------------
# Summary Update Functions
# --------------------------
def update_summary(file_path: str):
    """
    Update the summary file (SUMMARYPRO.txt) for a single Excel file.
    (This is the "summary modification event" for one file.)
    """
    try:
        data = extract_excel_data(file_path)
        if not data:
            return

        new_line = (
            f"{data['reference']}|{data['lieu']}|{data['charges_non_justifiees']}|"
            f"{data['main_oeuvre']}|{data['total_bons']}|{data['total_fournisseur']}|"
            f"{data['tva']}|{data['ttc']}|{data['ht']}|{data['benefice']}|"
            f"{data['tva_10']}|{data['etat_payment']}"
        )

        # Read existing summary lines (if any)
        if os.path.exists(SUMMARY_FILE):
            with open(SUMMARY_FILE, 'r') as f:
                lines = f.readlines()
        else:
            lines = []

        header = ("reference|lieu|TOTAL DES CHARGES NON JUSTIFIÉES|main_oeuvre|total_bons|"
                  "total_fournisseur|tva|ttc|ht|benefice|tva_10|etat_payment\n")
        if not lines or not lines[0].startswith("reference"):
            lines.insert(0, header)

        updated = False
        new_lines = [lines[0]]  # keep header
        for line in lines[1:]:
            if line.startswith(data['reference'] + '|'):
                new_lines.append(new_line + '\n')
                updated = True
            else:
                new_lines.append(line)
        if not updated:
            new_lines.append(new_line + '\n')

        with open(SUMMARY_FILE, 'w') as f:
            f.writelines(new_lines)

        # Update hash file
        current_hash = get_excel_hash(file_path)
        if os.path.exists(HASH_FILE):
            with open(HASH_FILE, 'r+') as f:
                try:
                    hashes = json.load(f)
                except json.JSONDecodeError:
                    hashes = {}
                hashes[file_path] = current_hash
                f.seek(0)
                json.dump(hashes, f)
                f.truncate()
        else:
            with open(HASH_FILE, 'w') as f:
                json.dump({file_path: current_hash}, f)

        logger.info(f"Updated summary for: {file_path}")
    except Exception as e:
        logger.error(f"Summary update error for {file_path}: {str(e)}")

def generate_summary():
    """
    Counting event: Scan all Excel files and rebuild the summary file.
    This ensures every Excel file has exactly one summary entry and
    orphaned or duplicate entries are removed.
    """
    try:
        excel_files = get_all_excel_files()
        header = ("reference|lieu|TOTAL DES CHARGES NON JUSTIFIÉES|main_oeuvre|total_bons|"
                  "total_fournisseur|tva|ttc|ht|benefice|tva_10|etat_payment\n")
        summary_lines = []
        hash_dict = {}

        for file_path in excel_files:
            data = extract_excel_data(file_path)
            if data:
                new_line = (
                    f"{data['reference']}|{data['lieu']}|{data['charges_non_justifiees']}|"
                    f"{data['main_oeuvre']}|{data['total_bons']}|{data['total_fournisseur']}|"
                    f"{data['tva']}|{data['ttc']}|{data['ht']}|{data['benefice']}|"
                    f"{data['tva_10']}|{data['etat_payment']}"
                )
                summary_lines.append(new_line)
                hash_dict[file_path] = get_excel_hash(file_path)

        with open(SUMMARY_FILE, 'w') as f:
            f.write(header)
            for line in summary_lines:
                f.write(line + '\n')

        with open(HASH_FILE, 'w') as f:
            json.dump(hash_dict, f)

        logger.info("Generated summary for all Excel files.")
    except Exception as e:
        logger.error(f"Error generating summary: {str(e)}")

# --------------------------
# File Monitoring Handlers
# --------------------------
class ExcelFileHandler(FileSystemEventHandler):
    """Monitor Excel files for changes and trigger summary events."""
    def on_created(self, event):
        if not event.is_directory and event.src_path.endswith(".xlsx"):
            logger.info(f"Detected creation of: {event.src_path}")
            time.sleep(2)  # Wait for the file to be fully saved
            update_summary(event.src_path)

    def on_modified(self, event):
        if not event.is_directory and event.src_path.endswith(".xlsx"):
            logger.info(f"Detected modification in: {event.src_path}")
            # Debounce modification events: reset timer if modifications occur in quick succession.
            if event.src_path in modification_timers:
                modification_timers[event.src_path].cancel()
            timer = threading.Timer(2.0, self.handle_modified, args=[event.src_path])
            modification_timers[event.src_path] = timer
            timer.start()

    def handle_modified(self, file_path):
        update_summary(file_path)
        if file_path in modification_timers:
            del modification_timers[file_path]

    def on_deleted(self, event):
        if not event.is_directory and event.src_path.endswith(".xlsx"):
            logger.info(f"Detected deletion of: {event.src_path}")
            # Trigger the counting event to rebuild the summary file
            generate_summary()

# --------------------------
# Start Monitoring
# --------------------------
def start_file_monitoring():
    """Begin monitoring Excel files for changes."""
    observer = Observer()
    event_handler = ExcelFileHandler()
    observer.schedule(event_handler, path=CHARGES_DIR, recursive=True)
    observer.start()
    logger.info(f"Started monitoring Excel files in: {CHARGES_DIR}")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()

# --------------------------
# Main Execution
# --------------------------
if __name__ == "__main__":
    # Ensure the summary file exists at startup
    initialize_summary_file()

    # Optionally, trigger a full generation (counting event) from the backend:
    # import sys
    # if len(sys.argv) > 1 and sys.argv[1] == "generate":
    #     generate_summary()
    #     exit(0)

    # Start monitoring Excel files
    start_file_monitoring()
